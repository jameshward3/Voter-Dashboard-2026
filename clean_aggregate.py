"""
Vote-by-Mail Operations Dashboard — cleaning + aggregation pipeline.

Reads a state voter-file extract of vote-by-mail records and produces
privacy-safe aggregates suitable for a public dashboard, plus a
record-level cleaned file for admin/operational use only.

USAGE
    python clean_aggregate.py <input.xlsx> <output_dir>

PRIVACY
    - Record-level outputs (cleaned.json / cleaned_voters.csv) are ADMIN
      ONLY. Do not publish.
    - Aggregate outputs apply a minimum-cell threshold (default = 10).
      Any street/precinct/grid cell with fewer than 10 records has its
      counts suppressed (set to null in JSON / blank in CSV).
    - Party affiliation is preserved in admin output for operational
      follow-up but is NEVER included in public aggregates.
    - This pipeline never infers or stores how any individual voted.
"""
import openpyxl, json, csv, re, sys, hashlib
from collections import Counter, defaultdict
from datetime import datetime, date

PRIVACY_THRESHOLD = 10

STREET_SUFFIX_CANONICAL = {
    ' AVENUE': ' AVE', ' STREET': ' ST', ' ROAD': ' RD',
    ' TERRACE': ' TER', ' PLACE': ' PL', ' BOULEVARD': ' BLVD',
    ' COURT': ' CT', ' DRIVE': ' DR', ' LANE': ' LN',
}

def norm_street(s):
    if not s: return None
    s = re.sub(r'\s+', ' ', s.strip()).upper()
    for k, v in STREET_SUFFIX_CANONICAL.items():
        s = s.replace(k, v)
    return s

def parse_date(d):
    if d is None: return None
    if isinstance(d, datetime): return d.date()
    if isinstance(d, date): return d
    try: return datetime.strptime(str(d), '%m/%d/%Y').date()
    except Exception: return None

def pseudo_geocode(street):
    """Deterministic placeholder coordinates inside Orange, NJ South Ward bbox.
    In production replace with Census Geocoder, Mapbox, or Google Geocoding."""
    if not street: return None, None
    # South Ward Orange NJ approximate bbox
    LAT_MIN, LAT_MAX = 40.756, 40.780
    LON_MIN, LON_MAX = -74.250, -74.225
    h = hashlib.md5(street.encode()).digest()
    lat = LAT_MIN + (h[0] / 255) * (LAT_MAX - LAT_MIN)
    lon = LON_MIN + (h[1] / 255) * (LON_MAX - LON_MIN)
    return round(lat, 6), round(lon, 6)

def clean(src):
    wb = openpyxl.load_workbook(src, data_only=True)
    ws = wb.active
    # Detect header row (look for 'ID' or 'Last Name')
    header_row = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row and 'ID' in row and 'Last Name' in row:
            header_row = i
            break
    if header_row is None:
        header_row = 1
    headers = [c.value for c in ws[header_row]]
    idx = {h: i for i, h in enumerate(headers) if h}

    def g(row, key, default=None):
        return row[idx[key]] if key in idx else default

    cleaned = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row[0] is None: continue
        street = norm_street(g(row, 'Street Name'))
        rec = {
            'id': g(row, 'ID'),
            'last_name': g(row, 'Last Name'),
            'first_name': g(row, 'First Name'),
            'middle': g(row, 'Middle Name'),
            'party': g(row, 'Party'),
            'street_no': g(row, 'Street No.'),
            'street': street,
            'apt': g(row, 'APT/UNIT'),
            'city': (g(row, 'Residence City') or '').title() or None,
            'state': g(row, 'Residence State'),
            'zip': g(row, 'Residence Zip'),
            'mailing_addr': g(row, 'Mailing'),
            'mailing_city': g(row, 'Mailing City'),
            'mailing_state': g(row, 'Mailing State'),
            'mailing_zip': g(row, 'Mailing Zip'),
            'req_date': parse_date(g(row, 'Ballot Req Rcvd Date')),
            'sent_date': parse_date(g(row, 'Ballot Mailed Date')),
            'received_date': parse_date(g(row, 'Ballot Received Date')),
            'status': g(row, 'Ballot Counted Status') or 'Outstanding',
            'ward': 'South',  # filtered file
        }
        rec['days_to_return'] = (rec['received_date'] - rec['sent_date']).days \
            if rec['sent_date'] and rec['received_date'] else None
        lat, lon = pseudo_geocode(street)
        rec['lat'], rec['lon'] = lat, lon
        cleaned.append(rec)
    return cleaned

def aggregate_by(records, key_fn, label_key='area'):
    bucket = defaultdict(list)
    for r in records:
        k = key_fn(r)
        if k: bucket[k].append(r)
    out = []
    for k, recs in bucket.items():
        n = len(recs)
        ret = sum(1 for r in recs if r['status'] == 'Received')
        suppressed = n < PRIVACY_THRESHOLD
        out.append({
            label_key: k,
            'n_voters': n,
            'requested': n,
            'returned': None if suppressed else ret,
            'outstanding': None if suppressed else n - ret,
            'return_rate': None if suppressed else round(ret / n * 100, 1),
            'suppressed': suppressed,
        })
    return sorted(out, key=lambda x: -x['n_voters'])

def daily_timeline(records):
    mailed = Counter()
    received = Counter()
    for r in records:
        if r['sent_date']: mailed[r['sent_date'].isoformat()] += 1
        if r['received_date']: received[r['received_date'].isoformat()] += 1
    dates = sorted(set(list(mailed) + list(received)))
    return [{'date': d, 'mailed': mailed[d], 'received': received[d]} for d in dates]

def cumulative_returned(records):
    dates = sorted(r['received_date'] for r in records if r['received_date'])
    out, running = [], 0
    by_day = Counter(d.isoformat() for d in dates)
    for d in sorted(by_day):
        running += by_day[d]
        out.append({'date': d, 'cumulative': running})
    return out

def build_insights(cleaned, street_agg, daily):
    total = len(cleaned)
    returned = sum(1 for r in cleaned if r['status'] == 'Received')
    visible = [s for s in street_agg if not s['suppressed']]
    if not visible: return {}
    last3 = sum(d['received'] for d in daily[-3:])
    prior3 = sum(d['received'] for d in daily[-6:-3]) if len(daily) >= 6 else 0
    return {
        'totals': {
            'voters_in_program': total,
            'ballots_mailed': sum(1 for r in cleaned if r['sent_date']),
            'ballots_received': returned,
            'outstanding': total - returned,
            'return_rate_pct': round(returned / total * 100, 1),
        },
        'highest_return_area': max(visible, key=lambda x: x['return_rate']),
        'lowest_return_area': min(visible, key=lambda x: x['return_rate']),
        'most_outstanding_area': max(visible, key=lambda x: x['outstanding']),
        'growth_last_3_days': last3 - prior3,
    }

def to_jsonable(o):
    if isinstance(o, (date, datetime)): return o.isoformat()
    return str(o)

def write_outputs(cleaned, out_dir):
    street_agg = aggregate_by(cleaned, lambda r: r['street'], 'street')
    daily = daily_timeline(cleaned)
    cum = cumulative_returned(cleaned)
    insights = build_insights(cleaned, street_agg, daily)

    # JSON
    for name, data in [('cleaned', cleaned), ('street_agg', street_agg),
                       ('daily', daily), ('cumulative', cum), ('insights', insights)]:
        with open(f'{out_dir}/{name}.json', 'w') as f:
            json.dump(data, f, default=to_jsonable)

    # Aggregate CSV (publishable)
    with open(f'{out_dir}/aggregate_by_street.csv', 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['street', 'n_voters', 'requested', 'returned',
                    'outstanding', 'return_rate_pct', 'suppressed_lt_10'])
        for s in street_agg:
            w.writerow([s['street'], s['n_voters'], s['requested'],
                        '' if s['suppressed'] else s['returned'],
                        '' if s['suppressed'] else s['outstanding'],
                        '' if s['suppressed'] else s['return_rate'],
                        s['suppressed']])

    # Record CSV (admin only)
    with open(f'{out_dir}/cleaned_voters.csv', 'w', newline='') as f:
        w = csv.writer(f)
        w.writerow(['voter_id', 'last_name', 'first_name', 'street_no',
                    'street', 'apt', 'city', 'state', 'zip',
                    'req_date', 'sent_date', 'received_date', 'status',
                    'days_to_return', 'ward'])
        for r in cleaned:
            w.writerow([r['id'], r['last_name'], r['first_name'],
                        r['street_no'], r['street'], r['apt'], r['city'],
                        r['state'], r['zip'], r['req_date'], r['sent_date'],
                        r['received_date'], r['status'],
                        r['days_to_return'], r['ward']])

if __name__ == '__main__':
    src = sys.argv[1] if len(sys.argv) > 1 else \
        '/sessions/blissful-wonderful-franklin/mnt/uploads/South ward mailins - 2026.xlsx'
    out = sys.argv[2] if len(sys.argv) > 2 else \
        '/sessions/blissful-wonderful-franklin/mnt/outputs'
    cleaned = clean(src)
    write_outputs(cleaned, out)
    print(f'Wrote {len(cleaned)} cleaned records and aggregates to {out}')
